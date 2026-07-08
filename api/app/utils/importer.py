import csv
import io
from typing import Any, Dict, List
import openpyxl


def parse_csv(file_bytes: bytes) -> List[Dict[str, Any]]:
    content = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    return [row for row in reader]


def parse_xlsx(file_bytes: bytes, sheet_name: str | None = None) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data = []
    for row in rows[1:]:
        row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers) and headers[i]}
        data.append(row_dict)
    
    return data


def process_import(
    file_bytes: bytes,
    filename: str,
    mapping: Dict[str, str],
    sheet_name: str | None = None
) -> List[Dict[str, Any]]:
    """
    Generic importer utility.
    :param file_bytes: raw bytes of the file
    :param filename: to determine extension (.csv or .xlsx)
    :param mapping: dict of {target_field: source_column_name}
    :param sheet_name: optional sheet name for xlsx
    :return: List of dictionaries with target fields
    """
    if filename.lower().endswith(".csv"):
        raw_data = parse_csv(file_bytes)
    elif filename.lower().endswith(".xlsx"):
        raw_data = parse_xlsx(file_bytes, sheet_name)
    else:
        raise ValueError("Unsupported file format. Must be .csv or .xlsx")

    results = []
    for row in raw_data:
        mapped_row = {}
        is_empty = True
        for target, source in mapping.items():
            val = row.get(source)
            if val is not None and str(val).strip() != "":
                is_empty = False
            mapped_row[target] = val
        
        # Skip completely empty rows
        if not is_empty:
            results.append(mapped_row)

    return results
